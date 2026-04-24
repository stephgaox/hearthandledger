"""
Universal bank CSV / Excel parser — no hardcoded bank formats.

Strategy:
  1. Scan the header row to identify column roles (date, description, amount, category).
  2. Sample data rows to detect the amount format (signed number, +/- prefix, debit/credit split).
  3. Parse every row using the detected layout.
  4. Map raw categories to our standard set.

Supports any bank that exports CSV, including Chase, Amex, Citi, Discover, PNC,
Wells Fargo, Capital One, Ally, etc. — without any bank-specific code.
"""

import os
import re
from datetime import date as date_type, datetime
from typing import Optional

# ── Standard category mapping ─────────────────────────────────────────────────

CATEGORY_MAP = {
    "annual fee":              "Other",
    "bill":                    "Bills & Utilities",
    "bills":                   "Bills & Utilities",
    "utilities":               "Bills & Utilities",
    "telecommunications":      "Bills & Utilities",
    "insurance":               "Bills & Utilities",
    "car":                     "Car",
    "automotive":              "Car",
    "gas":                     "Car",
    "gas stations":            "Car",
    "parking":                 "Travel",
    "rideshare":               "Travel",
    "cloth":                   "Shopping",
    "clothing":                "Shopping",
    "glothing":                "Shopping",
    "merchandise":             "Shopping",
    "shopping":                "Shopping",
    "electronics":             "Shopping",
    "furniture":               "Home",
    "home improvement":        "Home",
    "home":                    "Home",
    "daycare":                 "Kids & Childcare",
    "kid":                     "Kids & Childcare",
    "kids":                    "Kids & Childcare",
    "kindercare":              "Kids & Childcare",
    "education":               "Education",
    "books":                   "Education",
    "entertainment":           "Entertainment",
    "fast food":               "Food & Dining",
    "food and drink":          "Food & Dining",
    "food & drink":            "Food & Dining",
    "restaurants":             "Food & Dining",
    "dining":                  "Food & Dining",
    "groceries":               "Groceries",
    "grocery":                 "Groceries",
    "supermarkets":            "Groceries",
    "health & wellness":       "Medical",
    "medical":                 "Medical",
    "med":                     "Medical",
    "medication":              "Medical",
    "lego":                    "Entertainment",
    "sports":                  "Entertainment",
    "movies":                  "Entertainment",
    "music":                   "Entertainment",
    "pet":                     "Pet",
    "pet supplies":            "Pet",
    "refund":                  "Refund",
    "return":                  "Refund",
    "subscription":            "Subscriptions",
    "subscriptions":           "Subscriptions",
    "target":                  "Shopping",
    "travel":                  "Travel",
    "vacation":                "Travel",
    "airlines":                "Travel",
    "hotels":                  "Travel",
    "personal":                "Other",
    "professional services":   "Other",
    "fees & adjustments":      "Other",
    "other expenses":          "Other",
    "charity":                 "Other",
    "paychecks":               "Income",
    "payroll":                 "Income",
    "direct deposit":          "Income",
    "interest":                "Income",
}

SKIP_CATEGORIES_EXACT: set[str] = set()  # no rows are skipped — all transactions are shown

# Ordered keyword → category pairs for substring matching on the raw category string.
# Earlier entries take priority; more specific terms come first.
CATEGORY_MAPPING_KEYWORDS = [
    # Groceries (before generic "food")
    ("grocery",           "Groceries"),
    ("groceries",         "Groceries"),
    ("supermarket",       "Groceries"),
    ("wholesale store",   "Groceries"),
    # Food & Dining
    ("restaurant",        "Food & Dining"),
    ("dining",            "Food & Dining"),
    ("fast food",         "Food & Dining"),
    ("food & drink",      "Food & Dining"),
    ("food and drink",    "Food & Dining"),
    ("bar & caf",         "Food & Dining"),
    ("bakery",            "Food & Dining"),
    # Car (taxi/parking before "travel")
    ("taxi",              "Car"),
    ("car service",       "Car"),
    ("car services",      "Car"),
    ("rideshare",         "Travel"),
    ("parking",           "Travel"),
    ("gas station",       "Car"),
    ("gasoline",          "Car"),
    ("automotive",        "Car"),
    ("car rental",        "Car"),
    ("transit",           "Car"),
    ("toll",              "Car"),
    # Travel
    ("airline",           "Travel"),
    ("hotel",             "Travel"),
    ("motel",             "Travel"),
    ("vacation",          "Travel"),
    ("travel",            "Travel"),
    ("cruise",            "Travel"),
    # Medical / Healthcare
    ("healthcare",        "Medical"),
    ("medical",           "Medical"),
    ("pharmacy",          "Medical"),
    ("drug store",        "Medical"),
    ("dental",            "Medical"),
    ("vision",            "Medical"),
    ("optom",             "Medical"),
    # Bills & Utilities
    ("utilities",         "Bills & Utilities"),
    ("insurance",         "Bills & Utilities"),
    ("telecom",           "Bills & Utilities"),
    ("telephone",         "Bills & Utilities"),
    ("wireless",          "Bills & Utilities"),
    # Subscriptions (before entertainment)
    ("streaming",         "Subscriptions"),
    ("subscription",      "Subscriptions"),
    # Entertainment
    ("entertainment",     "Entertainment"),
    ("sport",             "Entertainment"),
    ("movie",             "Entertainment"),
    ("theater",           "Entertainment"),
    ("theatre",           "Entertainment"),
    ("amusement",         "Entertainment"),
    # Kids & Childcare
    ("child",             "Kids & Childcare"),
    ("daycare",           "Kids & Childcare"),
    ("day care",          "Kids & Childcare"),
    # Education
    ("education",         "Education"),
    ("school",            "Education"),
    ("tuition",           "Education"),
    ("book",              "Education"),
    # Home
    ("home improvement",  "Home"),
    ("hardware",          "Home"),
    ("furniture",         "Home"),
    # Shopping (broad — comes after specific categories)
    ("merchandise",       "Shopping"),
    ("shopping",          "Shopping"),
    ("retail",            "Shopping"),
    ("clothing",          "Shopping"),
    ("apparel",           "Shopping"),
    ("electronics",       "Shopping"),
    # Pet
    ("pet",               "Pet"),
    ("veterinarian",      "Pet"),
    # Income / Refund
    ("paycheck",          "Income"),
    ("payroll",           "Income"),
    ("refund",            "Refund"),
]

# Common merchant name fragments → category, applied to the transaction description
# when the category string is still unresolved.
MERCHANT_KEYWORDS = [
    # Groceries
    ("whole foods",    "Groceries"),
    ("trader joe",     "Groceries"),
    ("kroger",         "Groceries"),
    ("safeway",        "Groceries"),
    ("publix",         "Groceries"),
    ("aldi",           "Groceries"),
    ("costco",         "Groceries"),
    ("sam's club",     "Groceries"),
    ("bj's",           "Groceries"),
    ("wegmans",        "Groceries"),
    ("sprouts",        "Groceries"),
    # Food & Dining
    ("mcdonald",       "Food & Dining"),
    ("starbucks",      "Food & Dining"),
    ("chipotle",       "Food & Dining"),
    ("subway",         "Food & Dining"),
    ("dunkin",         "Food & Dining"),
    ("doordash",       "Food & Dining"),
    ("uber eats",      "Food & Dining"),
    ("grubhub",        "Food & Dining"),
    ("panera",         "Food & Dining"),
    ("chick-fil-a",    "Food & Dining"),
    ("domino",         "Food & Dining"),
    # Car
    ("uber",           "Travel"),
    ("lyft",           "Travel"),
    ("ez pass",        "Travel"),
    ("e-zpass",        "Travel"),
    ("sunoco",         "Car"),
    ("exxon",          "Car"),
    ("chevron",        "Car"),
    ("speedway",       "Car"),
    # Travel
    ("delta",          "Travel"),
    ("united air",     "Travel"),
    ("american air",   "Travel"),
    ("southwest air",  "Travel"),
    ("jetblue",        "Travel"),
    ("marriott",       "Travel"),
    ("hilton",         "Travel"),
    ("hyatt",          "Travel"),
    ("airbnb",         "Travel"),
    ("expedia",        "Travel"),
    # Subscriptions
    ("netflix",        "Subscriptions"),
    ("spotify",        "Subscriptions"),
    ("hulu",           "Subscriptions"),
    ("disney+",        "Subscriptions"),
    ("apple.com/bill", "Subscriptions"),
    ("amazon prime",   "Subscriptions"),
    ("youtube",        "Subscriptions"),
    ("hbo",            "Subscriptions"),
    ("paramount",      "Subscriptions"),
    ("peacock",        "Subscriptions"),
    # Shopping / Home
    ("amazon",         "Shopping"),
    ("target",         "Shopping"),
    ("walmart",        "Shopping"),
    ("best buy",       "Shopping"),
    ("home depot",     "Home"),
    ("lowe's",         "Home"),
    ("ikea",           "Home"),
    ("wayfair",        "Home"),
    # Medical / Health
    ("cvs",            "Medical"),
    ("walgreens",      "Medical"),
    ("rite aid",       "Medical"),
    ("upmc",           "Medical"),
    ("kaiser",         "Medical"),
    ("aetna",          "Medical"),
    ("cigna",          "Medical"),
    ("pharmacy",       "Medical"),
    ("hospital",       "Medical"),
    ("clinic",         "Medical"),
    ("dental",         "Medical"),
    ("vision",         "Medical"),
    ("optometrist",    "Medical"),
    ("urgent care",    "Medical"),
    # Pet
    ("petco",          "Pet"),
    ("petsmart",       "Pet"),
    ("chewy",          "Pet"),
    ("rover.com",      "Pet"),
    ("rover",          "Pet"),          # Rover pet sitting/walking
    # Entertainment
    ("ticketmaster",   "Entertainment"),
    ("stubhub",        "Entertainment"),
    ("eventbrite",     "Entertainment"),
    ("amc theatre",    "Entertainment"),
    ("regal",          "Entertainment"),
    ("cinemark",       "Entertainment"),
    ("ppg paints arena","Entertainment"),
    ("arena",          "Entertainment"),
    ("museum",         "Entertainment"),
    ("carnegie",       "Entertainment"),
    ("arts",           "Entertainment"),
    # Food & Dining (more chains)
    ("instacart",      "Food & Dining"),
    ("grubhub",        "Food & Dining"),
    ("tst*",           "Food & Dining"),   # Toast POS prefix
    ("sq *",           "Food & Dining"),   # Square POS prefix (restaurants)
    ("le pain",        "Food & Dining"),
    ("sushi",          "Food & Dining"),
    ("ramen",          "Food & Dining"),
    ("pizza",          "Food & Dining"),
    ("burrito",        "Food & Dining"),
    ("cafe",           "Food & Dining"),
    ("bakery",         "Food & Dining"),
    ("diner",          "Food & Dining"),
    ("grill",          "Food & Dining"),
    ("bistro",         "Food & Dining"),
    ("restaurant",     "Food & Dining"),
    ("eatery",         "Food & Dining"),
    # Groceries
    ("giant eagle",    "Groceries"),
    ("harris teeter",  "Groceries"),
    ("meijer",         "Groceries"),
    ("food lion",      "Groceries"),
    ("oriental market","Groceries"),
    ("asian market",   "Groceries"),
    ("sprout",         "Groceries"),    # Sprouts Farmers Market (also "SP SPROUT" on Amex)
    # Shopping / Retail
    ("nordstrom",      "Shopping"),
    ("macy",           "Shopping"),
    ("saks",           "Shopping"),
    ("sephora",        "Shopping"),
    ("uniqlo",         "Shopping"),
    ("bed bath",       "Shopping"),
    ("gap",            "Shopping"),
    ("old navy",       "Shopping"),
    ("h&m",            "Shopping"),
    ("zara",           "Shopping"),
    ("etsy",           "Shopping"),
    ("lululemon",      "Shopping"),
    ("nike",           "Shopping"),
    ("adidas",         "Shopping"),
    ("under armour",   "Shopping"),
    ("tj maxx",        "Shopping"),
    ("marshalls",      "Shopping"),
    # Subscriptions (streaming/services)
    ("disneyplus",     "Subscriptions"),
    ("disney plus",    "Subscriptions"),   # Amex writes with a space
    ("disney+",        "Subscriptions"),
    ("openai",         "Subscriptions"),
    ("cursor",         "Subscriptions"),
    ("whoop",          "Subscriptions"),
    ("ringcentral",    "Subscriptions"),
    ("adobe",          "Subscriptions"),
    ("microsoft",      "Subscriptions"),
    ("dropbox",        "Subscriptions"),
    ("google",         "Subscriptions"),
    # Bills & Utilities
    ("comcast",        "Bills & Utilities"),
    ("xfinity",        "Bills & Utilities"),
    ("verizon",        "Bills & Utilities"),
    ("at&t",           "Bills & Utilities"),
    ("t-mobile",       "Bills & Utilities"),
    ("electric",       "Bills & Utilities"),
    ("duquesne",       "Bills & Utilities"),
    ("peco",           "Bills & Utilities"),
    ("water",          "Bills & Utilities"),
    ("hampton water",  "Bills & Utilities"),
    ("govpmt",         "Bills & Utilities"),  # government payments
    # Travel
    ("enterprise rent","Travel"),
    ("hertz",          "Travel"),
    ("avis",           "Travel"),
    ("budget rent",    "Travel"),
    ("amex fine hotel","Travel"),
    ("amextravel",     "Travel"),
    ("amex travel",    "Travel"),
    ("american express tra", "Travel"),   # Amex travel booking prefix
    ("parkwhiz",       "Travel"),
    ("spothero",       "Travel"),
    # Kids & Childcare
    ("shadyside academy", "Kids & Childcare"),
    ("kindercare",     "Kids & Childcare"),
    ("activeworks",    "Kids & Childcare"),  # camp/activity registration
    # Home
    ("pest",           "Home"),
    ("ridgewood",      "Home"),
    ("plumber",        "Home"),
    ("electrician",    "Home"),

    # ── Semantic category indicators ─────────────────────────────────────────
    # These are generic words that commonly appear *within* transaction
    # descriptions and signal a category, even for unknown merchants.
    # Kept last so specific merchant names above always win.

    # Travel
    ("hotel",          "Travel"),
    ("hotels",         "Travel"),
    ("resort",         "Travel"),
    ("motel",          "Travel"),
    ("lodging",        "Travel"),
    ("travel",         "Travel"),
    ("airline",        "Travel"),
    ("airways",        "Travel"),
    ("cruise",         "Travel"),
    ("rental car",     "Travel"),
    ("car rental",     "Travel"),
    ("parking",        "Travel"),

    # Food & Dining
    ("dining",         "Food & Dining"),
    ("noodl",          "Food & Dining"),    # "Noodletalk", "Noodle House", etc.
    ("burger",         "Food & Dining"),
    ("taco",           "Food & Dining"),
    ("wing",           "Food & Dining"),
    ("bbq",            "Food & Dining"),
    ("boba",           "Food & Dining"),

    # Groceries
    ("grocery",        "Groceries"),
    ("groceries",      "Groceries"),
    ("supermarket",    "Groceries"),
    ("farmers market", "Groceries"),

    # Bills & Utilities
    ("utility",        "Bills & Utilities"),  # "UTILITYPMT", "utility payment"
    ("utilities",      "Bills & Utilities"),
    ("internet",       "Bills & Utilities"),
    ("cable tv",       "Bills & Utilities"),
    ("insurance",      "Bills & Utilities"),
    ("tax service",    "Bills & Utilities"),  # "JORDANTAXSERVICE"

    # Medical
    ("medical",        "Medical"),
    ("health",         "Medical"),
    ("doctor",         "Medical"),
    ("physician",      "Medical"),
    ("orthopedic",     "Medical"),
    ("surgery",        "Medical"),
    ("veterinary",     "Pet"),              # vet before generic "health"
    ("animal hosp",    "Pet"),

    # Car
    ("gas station",    "Car"),
    ("car wash",       "Car"),
    ("auto repair",    "Car"),
    ("auto service",   "Car"),
    ("tire",           "Car"),
    ("oil change",     "Car"),
    ("smog",           "Car"),

    # Home
    ("hardware",       "Home"),
    ("furniture",      "Home"),
    ("flooring",       "Home"),
    ("roofing",        "Home"),
    ("storage",        "Home"),
    ("moving",         "Home"),

    # Shopping
    ("outlet",         "Shopping"),
    ("retail",         "Shopping"),
    ("boutique",       "Shopping"),

    # Subscriptions
    ("membership",     "Subscriptions"),
    ("subscription",   "Subscriptions"),
    ("streaming",      "Subscriptions"),

    # Kids & Childcare
    ("daycare",        "Kids & Childcare"),
    ("preschool",      "Kids & Childcare"),
    ("tutoring",       "Kids & Childcare"),
    ("childcare",      "Kids & Childcare"),
    ("summer camp",    "Kids & Childcare"),

    # Pet
    ("pet supply",     "Pet"),
    ("pet store",      "Pet"),
    ("pet sitting",    "Pet"),
]


def _map_category(raw: str, description: str = "") -> str:
    s = raw.strip().lower()

    # Pass 1: exact dict match (fastest)
    if s and s in CATEGORY_MAP:
        return CATEGORY_MAP[s]

    # Pass 2: ordered keyword substring scan on category string
    if s:
        for keyword, category in CATEGORY_MAPPING_KEYWORDS:
            if keyword in s:
                return category

    # Pass 3: description-based merchant matching
    desc = description.strip().lower()
    if desc:
        for keyword, category in MERCHANT_KEYWORDS:
            if keyword in desc:
                return category

    return "Other"




# ── Date parsing ──────────────────────────────────────────────────────────────

DATE_FORMATS = [
    "%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y", "%d/%m/%Y",
    "%m-%d-%Y", "%Y/%m/%d", "%b %d, %Y", "%d %b %Y",
]


def _parse_date(value) -> Optional[date_type]:
    if isinstance(value, date_type) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        s = value.strip().strip('"')
        # Skip "PENDING - ..." style dates
        if s.upper().startswith("PENDING"):
            return None
        for fmt in DATE_FORMATS:
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                pass
    return None


def _looks_like_date(value: str) -> bool:
    return _parse_date(value) is not None


# ── Amount parsing ────────────────────────────────────────────────────────────

def _clean_amount(raw: str) -> Optional[float]:
    """
    Parse amount strings in any format banks use:
      - "- $1,234.56"  → -1234.56  (PNC style: sign + space + dollar)
      - "+$500"        → 500.0
      - "(123.45)"     → -123.45   (accounting negative)
      - "-50.00"       → -50.0     (simple signed)
      - "1,234.56"     → 1234.56   (no sign = positive)
    """
    s = raw.strip().strip('"').strip()
    if not s:
        return None

    negative = False

    # Parentheses = negative (accounting format)
    if s.startswith("(") and s.endswith(")"):
        negative = True
        s = s[1:-1]

    # Leading "- " or "+" with optional space
    if s.startswith("-"):
        negative = True
        s = s[1:].strip()
    elif s.startswith("+"):
        s = s[1:].strip()

    # Remove currency symbols and commas
    s = s.replace("$", "").replace(",", "").replace(" ", "")

    try:
        amount = float(s)
        return -amount if negative else amount
    except ValueError:
        return None


# ── Column role detection ─────────────────────────────────────────────────────

# Keywords that identify each column role
DATE_KEYWORDS      = {"date", "time", "posted", "post", "transaction date", "trans. date",
                      "trans date", "posting date", "value date"}
# High-priority desc keywords are tried across ALL columns before low-priority ones.
# This prevents "Details" (Chase bank type indicator) from shadowing "Description".
DESC_KEYWORDS_HIGH = {"description", "desc", "memo", "payee", "merchant",
                      "transaction description", "narrative"}
DESC_KEYWORDS_LOW  = {"name", "details"}
AMOUNT_KEYWORDS    = {"amount", "transaction amount", "amt"}
DEBIT_KEYWORDS     = {"debit", "withdrawal", "charge", "spent"}
CREDIT_KEYWORDS    = {"credit", "deposit", "payment received"}
CATEGORY_KEYWORDS  = {"category", "type", "transaction type", "class"}


def _find_col(header: list[str], keywords: set) -> Optional[int]:
    """Return index of first header cell that matches any keyword (exact or substring)."""
    for i, h in enumerate(header):
        h_low = h.strip().lower()
        if h_low in keywords:
            return i
        if any(k in h_low for k in keywords):
            return i
    return None


def _find_desc_col(header: list[str]) -> Optional[int]:
    """Find description column, preferring high-confidence keywords over low-confidence ones.

    Chase bank exports have both 'Details' (type indicator: DEBIT/CREDIT) and
    'Description' (actual merchant name). Scanning left-to-right would stop at
    'Details', so we check high-priority keywords across all columns first.
    """
    result = _find_col(header, DESC_KEYWORDS_HIGH)
    if result is not None:
        return result
    return _find_col(header, DESC_KEYWORDS_LOW)


def _detect_columns(header: list[str], sample_rows: list[list[str]]) -> dict:
    """
    Returns a dict describing the column layout:
      date_col, desc_col, amount_col, debit_col, credit_col, cat_col,
      amount_format: "signed" | "debit_credit" | "prefix_sign"
    """
    h = [c.strip().lower() for c in header]

    date_col   = _find_col(h, DATE_KEYWORDS)
    desc_col   = _find_desc_col(h)
    amount_col = _find_col(h, AMOUNT_KEYWORDS)
    debit_col  = _find_col(h, DEBIT_KEYWORDS)
    credit_col = _find_col(h, CREDIT_KEYWORDS)
    cat_col    = _find_col(h, CATEGORY_KEYWORDS)

    # If no date column found by keyword, probe columns for parseable dates
    if date_col is None:
        for i in range(min(4, len(header))):
            hits = sum(
                1 for row in sample_rows[:10]
                if len(row) > i and _looks_like_date(row[i])
            )
            if hits >= 3:
                date_col = i
                break

    # Determine amount format
    positive_expense = False  # default: positive = income (Chase/Citi)

    if debit_col is not None and credit_col is not None:
        fmt = "debit_credit"
    elif amount_col is not None:
        # Sample to see if values use "- $x" prefix style
        prefix_hits = 0
        pos_count = 0
        neg_count = 0
        for row in sample_rows[:20]:
            if len(row) > amount_col:
                v = row[amount_col].strip().strip('"')
                if re.match(r'^[+\-]\s*\$', v):
                    prefix_hits += 1
                val = _clean_amount(v)
                if val is not None and val != 0:
                    if val > 0:
                        pos_count += 1
                    else:
                        neg_count += 1

        if prefix_hits >= 2:
            fmt = "prefix_sign"
        else:
            fmt = "signed"
            # If the majority of sampled amounts are positive, the file looks like
            # Amex/Discover convention: positive = charge (expense).
            # BUT first check if the positive rows are income-labelled (interest,
            # dividend, deposit, lending) — that means it's actually a bank/investment
            # account where positive = income, negative = expense.
            total = pos_count + neg_count
            if total > 0 and pos_count / total >= 0.65:
                income_desc_hits = 0
                if desc_col is not None:
                    for row in sample_rows[:20]:
                        if len(row) > amount_col and len(row) > desc_col:
                            val = _clean_amount(row[amount_col])
                            if val and val > 0:
                                d = row[desc_col].strip().lower()
                                if any(k in d for k in (
                                    "interest", "dividend", "deposit",
                                    "lending", "payroll", "paycheck",
                                )):
                                    income_desc_hits += 1
                # If 2+ positive rows describe income, treat as bank/investment style
                positive_expense = (income_desc_hits < 2)
            elif desc_col is not None:
                # Ratio < 65% but check for Amex-specific description patterns.
                # Amex Platinum cards have many statement credits (hotel credit,
                # entertainment credit, Walmart+ credit) that inflate negative counts,
                # pushing the ratio below 65% even though positive = expense convention.
                # "MOBILE PAYMENT - THANK YOU" appearing as a negative amount is a
                # definitive Amex signal: it's the payment confirmation Amex prints
                # when you pay your bill, which reduces balance (negative in Amex convention).
                amex_signals = 0
                for row in sample_rows[:20]:
                    if len(row) > amount_col and len(row) > desc_col:
                        val = _clean_amount(row[amount_col])
                        d = row[desc_col].strip().lower()
                        if val and val < 0 and any(k in d for k in (
                            "mobile payment",          # "MOBILE PAYMENT - THANK YOU"
                            "platinum hotel credit",
                            "platinum digital",
                            "platinum walmart",
                            "shop saks with platinum",
                            "amex epay",
                        )):
                            amex_signals += 1
                if amex_signals >= 1:
                    positive_expense = True
    else:
        fmt = "unknown"

    # Find a separate transaction-direction indicator column (e.g. Chase "Type": Sale/Return).
    # Only meaningful when it's a different column from cat_col.
    type_indicator_col = None
    for i, h_item in enumerate(h):
        if i == cat_col:
            continue
        if h_item.strip() == "type":
            type_indicator_col = i
            break

    return {
        "date_col":           date_col,
        "desc_col":           desc_col,
        "amount_col":         amount_col,
        "debit_col":          debit_col,
        "credit_col":         credit_col,
        "cat_col":            cat_col,
        "type_indicator_col": type_indicator_col,
        "amount_format":      fmt,
        "positive_expense":   positive_expense,
    }


# ── Universal CSV parser ──────────────────────────────────────────────────────

def _parse_universal(header: list[str], rows: list[list[str]]) -> list[dict]:
    cols = _detect_columns(header, rows)

    date_col   = cols["date_col"]
    desc_col   = cols["desc_col"]
    amount_col = cols["amount_col"]
    debit_col  = cols["debit_col"]
    credit_col = cols["credit_col"]
    cat_col           = cols["cat_col"]
    type_indicator_col = cols["type_indicator_col"]
    fmt               = cols["amount_format"]
    positive_expense  = cols["positive_expense"]

    if date_col is None or desc_col is None:
        raise ValueError(
            "Could not identify date and description columns. "
            f"Header: {header}"
        )

    transactions = []

    for row in rows:
        if not any(c.strip() for c in row):
            continue

        # Date
        date_val = row[date_col] if len(row) > date_col else ""
        tx_date = _parse_date(date_val)
        if tx_date is None:
            continue

        description = row[desc_col].strip().strip('"') if len(row) > desc_col else ""
        if not description:
            continue

        category_raw = row[cat_col].strip().strip('"') if cat_col is not None and len(row) > cat_col else ""

        # ── Amount & direction ────────────────────────────────────────────────
        amount: Optional[float] = None
        tx_type: Optional[str]  = None

        if fmt == "debit_credit":
            d = _clean_amount(row[debit_col]) if debit_col is not None and len(row) > debit_col else None
            c = _clean_amount(row[credit_col]) if credit_col is not None and len(row) > credit_col else None
            if d and abs(d) > 0:
                amount, tx_type = abs(d), "expense"
            elif c and abs(c) > 0:
                amount, tx_type = abs(c), "income"

        elif fmt in ("signed", "prefix_sign"):
            if amount_col is not None and len(row) > amount_col:
                raw_amount = _clean_amount(row[amount_col])
                if raw_amount is not None and raw_amount != 0:
                    amount = abs(raw_amount)
                    # Positive = money in vs money out depends on the bank:
                    # Chase: positive = credit/income, negative = expense
                    # Amex/Discover: positive = charge (expense), negative = credit
                    # PNC: explicit +/- prefix
                    if fmt == "prefix_sign":
                        # PNC-style: sign is explicit in the string
                        raw_str = row[amount_col].strip().strip('"')
                        tx_type = "income" if raw_str.lstrip().startswith("+") else "expense"
                    else:
                        # For signed: use category hint to decide direction
                        cat_low = category_raw.lower()
                        desc_low = description.lower()
                        if cat_low in {"paychecks", "payroll", "direct deposit",
                                       "interest", "refund", "credit"}:
                            tx_type = "income"
                        elif any(k in desc_low for k in (
                            "interest payment", "interest earned", "dividend",
                            "stock lending", "ach deposit", "direct deposit",
                            "payroll", "paycheck",
                        )):
                            # Description strongly implies income regardless of sign convention
                            tx_type = "income"
                        elif raw_amount > 0 and any(
                            k in (header[amount_col] if amount_col is not None else "").lower()
                            for k in ("charge", "debit", "spent")
                        ):
                            tx_type = "expense"
                        else:
                            raw_val = _clean_amount(row[amount_col])
                            if positive_expense:
                                # Amex/Discover: positive = charge, negative = credit
                                tx_type = "expense" if (raw_val or 0) > 0 else "income"
                            else:
                                # Chase/Citi: positive = income, negative = expense
                                tx_type = "income" if (raw_val or 0) > 0 else "expense"

        if amount is None or amount == 0:
            continue

        # Classify inter-account payments as transfers (not income/expense)
        desc_low = description.lower()
        cat_low_check = category_raw.strip().lower()

        # ── Classify inter-account transactions ───────────────────────────────
        # All detection flags defined upfront so category assignment is clean.

        _is_venmo_zelle = (
            "venmo" in desc_low
            or "zel to" in desc_low
            or "zel from" in desc_low
            or desc_low.startswith("zelle")
        )

        # Bank paying a CC bill (bank statement side):
        # PNC Category="Credit Card Payments", Chase Type="loan_pmt", or
        # description identifies the CC brand being paid (epay, autopay, etc.)
        _is_bank_cc_payment = not _is_venmo_zelle and (
            cat_low_check in {"credit card payments", "loan_pmt"}
            or any(k in desc_low for k in (
                "epay",              # "AMEX EPAYMENT ACH PMT", "CHASE CREDIT CRD EPAY"
                "autopay",
                "credit card pay",
                "card srvc payment", # "TARGET CARD SRVC PAYMENT"
                "card svc payment",
            ))
        )

        # CC statement receiving a payment from the cardholder:
        # Chase CC Type="Payment" → description "Payment Thank You-Mobile"
        # Amex: "MOBILE PAYMENT - THANK YOU" (negative = credit in Amex convention)
        _is_cc_received = not _is_venmo_zelle and not _is_bank_cc_payment and any(
            k in desc_low for k in (
                "payment thank you",  # Chase CC: "Payment Thank You-Mobile"
                "mobile payment",     # Amex: "MOBILE PAYMENT - THANK YOU"
                "payment received",
                "ach payment",
                "automatic payment",
                "internet payment",
                "statement payment",
                "online payment",
            )
        )

        # Plain bank-to-bank transfer
        _is_plain_transfer = not _is_venmo_zelle and not _is_bank_cc_payment and (
            cat_low_check in {"transfers", "acct_xfer"}
            or any(k in desc_low for k in (
                "online transfer",   # "ONLINE TRANSFER TO/FROM"
                "realtime transfer", # Chase inter-account
            ))
        )

        # Transfer to a known investment / brokerage platform (outgoing from bank)
        _is_investment_transfer = (
            not _is_venmo_zelle and not _is_bank_cc_payment and not _is_plain_transfer
            and any(k in desc_low for k in (
                "robinhood", "fidelity", "vanguard", "schwab", "etrade",
                "td ameritrade", "merrill", "ally invest", "betterment",
                "wealthfront", "acorns", "stash invest", "sofi invest",
                "jpmorgan invest", "goldman sachs",
            ))
        )

        # ── Override tx_type based on transaction class ───────────────────────
        # Save sign direction before any override (income=positive, expense=negative)
        _was_income = tx_type == "income"

        if _is_venmo_zelle:
            if any(k in desc_low for k in ("cashout", "cash out", "zel from", "ach credit")):
                tx_type = "income"
            else:
                tx_type = "expense"
        elif _is_bank_cc_payment:
            tx_type = "transfer_out"  # CC bill payment leaving the bank
        elif _is_cc_received:
            tx_type = "transfer_in"   # payment arriving at the CC account
        elif _is_plain_transfer:
            tx_type = "transfer_in" if _was_income else "transfer_out"
        elif _is_investment_transfer:
            tx_type = "transfer_out"  # money leaving bank to investment platform
        # else: keep tx_type from amount/sign detection

        # ── Category assignment ───────────────────────────────────────────────
        if _is_bank_cc_payment:
            category = "CC Payments"
        elif _is_cc_received:
            category = "Payment Received"
        elif tx_type in ("transfer_in", "transfer_out"):
            category = "Transfer"
        elif _is_venmo_zelle and tx_type == "income":
            category = "Income"
        elif _is_venmo_zelle and tx_type == "expense":
            category = "Food & Dining"  # best default for peer payments; user can recategorize
        elif tx_type == "income":
            # Always try description-based merchant matching even when category column is empty
            category = _map_category(category_raw, description)
            if category == "Other":
                category = "Income"
        else:
            # Always try description-based merchant matching (handles Amex — no category column)
            category = _map_category(category_raw, description)

        # Chase "Type" column: an explicit "Return" overrides sign-based classification.
        if type_indicator_col is not None and len(row) > type_indicator_col:
            if row[type_indicator_col].strip().lower() == "return":
                tx_type = "income"
                category = "Refund"

        transactions.append({
            "date":        tx_date.isoformat(),
            "description": description,
            "amount":      amount,
            "type":        tx_type,
            "category":    category,
            "account":     None,
        })

    return transactions


# ── Excel (Family Money Tracker format) ───────────────────────────────────────

def parse_excel(file_path: str) -> list[dict]:
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    expense_sheets = [s for s in wb.sheetnames if s.lower().startswith("expense")]

    transactions = []
    for sheet_name in expense_sheets:
        ws = wb[sheet_name]
        header_found = False

        for row in ws.iter_rows(values_only=True):
            if row[0] == "Time":
                header_found = True
                continue
            if not header_found:
                continue

            tx_date, account, spend, category, details, notes = (list(row) + [None] * 6)[:6]
            tx_date = _parse_date(tx_date)
            if tx_date is None or spend is None or spend == 0:
                continue

            cat_raw = str(category or "").strip()
            transactions.append({
                "date":        tx_date.isoformat(),
                "description": str(details or account or "Transaction").strip(),
                "amount":      abs(float(spend)),
                "type":        "expense",
                "category":    _map_category(cat_raw, str(details or account or "")),
                "account":     str(account or "").strip() or None,
            })

    if not transactions:
        raise ValueError(
            "No expense sheets found. "
            "Expected sheets named 'Expense - Jan', 'Expense - Feb', etc."
        )
    return transactions


# ── CSV entry point ───────────────────────────────────────────────────────────

def parse_csv(file_path: str) -> list[dict]:
    import csv

    with open(file_path, "r", encoding="utf-8-sig", errors="replace") as f:
        reader = csv.reader(f)
        all_rows = [r for r in reader if any(c.strip() for c in r)]

    if not all_rows:
        raise ValueError("CSV file appears to be empty.")

    return _parse_universal(all_rows[0], all_rows[1:])


# ── Main entry ────────────────────────────────────────────────────────────────

def parse_file_direct(file_path: str) -> list[dict]:
    ext = os.path.splitext(file_path)[1].lower()
    if ext in (".xlsx", ".xls"):
        return parse_excel(file_path)
    elif ext == ".csv":
        return parse_csv(file_path)
    else:
        raise ValueError(
            f"Direct parsing not supported for '{ext}'. Use AI upload instead."
        )
